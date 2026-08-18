from pathlib import Path
import subprocess
from openpyxl import load_workbook

base = Path('/home/ubuntu/balancerts-erp/docs')
text_dir = base / 'agt-portal-extracted'
text_dir.mkdir(exist_ok=True)

pdf_out = text_dir / 'agt-portal.pdf.txt'
subprocess.run(['pdftotext', '-layout', str(base / 'agt-portal.pdf'), str(pdf_out)], check=True)

for xlsx_name in ('agt-portal-1.xlsx', 'agt-portal-2.xlsx'):
    workbook = load_workbook(base / xlsx_name, data_only=False, read_only=True)
    out = text_dir / f'{xlsx_name}.txt'
    with out.open('w', encoding='utf-8') as handle:
        for sheet in workbook.worksheets:
            handle.write(f'## SHEET: {sheet.title}\n')
            for row in sheet.iter_rows(values_only=True):
                values = ['' if value is None else str(value) for value in row]
                handle.write('\t'.join(values).rstrip() + '\n')
            handle.write('\n')
print(f'PDF text: {pdf_out.stat().st_size} bytes')
for path in sorted(text_dir.glob('*.xlsx.txt')):
    print(f'{path.name}: {path.stat().st_size} bytes')
