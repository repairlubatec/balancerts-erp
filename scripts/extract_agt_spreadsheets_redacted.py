from pathlib import Path
import re
from openpyxl import load_workbook

base = Path('/home/ubuntu/balancerts-erp/docs')
output = base / 'agt-spreadsheets-redacted.md'
files = [base / 'agt-portal-1.xlsx', base / 'agt-portal-2.xlsx']
redacted = []
for path in files:
    workbook = load_workbook(path, data_only=False, read_only=True)
    redacted.append(f'## {path.name}')
    redacted.append(f'Folhas: {", ".join(workbook.sheetnames)}')
    for sheet in workbook.worksheets:
        redacted.append(f'### Folha: {sheet.title}')
        for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            values = []
            for value in row:
                text = '' if value is None else str(value)
                text = re.sub(r'-----BEGIN PRIVATE KEY-----.+?-----END PRIVATE KEY-----', '[PRIVATE KEY REDACTED]', text, flags=re.S)
                text = re.sub(r'-----BEGIN PUBLIC KEY-----.+?-----END PUBLIC KEY-----', '[PUBLIC KEY REDACTED]', text, flags=re.S)
                if len(text) > 300:
                    text = text[:120] + ' [LONG VALUE REDACTED]'
                values.append(text)
            if any(values):
                redacted.append(f'{row_index}: ' + ' | '.join(values))
        redacted.append('')
output.write_text('\n'.join(redacted) + '\n', encoding='utf-8')
print(output)
